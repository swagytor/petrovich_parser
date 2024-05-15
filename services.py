import pprint
import time
from math import ceil

from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, Alignment, Font
from selenium import webdriver
from selenium.common import StaleElementReferenceException, NoSuchElementException, ElementClickInterceptedException, \
    WebDriverException, SessionNotCreatedException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By

import settings
from google_services import download_file_from_google_drive, upload_file_to_google_drive

ITEMS_PER_PAGE = 20
BASE_URL = 'https://petrovich.ru'
thins = Side(border_style="thin", color="000000")
bold = Side(border_style="medium", color="000000")
thin_border = Border(left=thins, right=thins, top=thins, bottom=thins)
bold_border = Border(left=bold, right=bold, top=bold, bottom=bold)
bold_font = Font(bold=True)
align = Alignment(horizontal='center', vertical='center')


# def check_proxies(proxy):
#     print(f'Пробуем {proxy}')
#     proxy = proxy.strip()
#     try:
#         res = requests.get('https://petrovich.ru/', proxies={'http': proxy,
#                                                              'https': proxy}, )
#
#         if res.status_code == 200:
#             print(proxy)
#     except:
#         pass

def get_web_driver():
    ua = UserAgent(platforms='pc', os=['windows', 'macos', 'linux'])
    user_agent = ua.random
    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent={user_agent}')
    options.add_argument('--no-sandbox')
    options.add_argument('--log-level=3')
    # options.add_argument('--window-size=1920,1080')
    options.add_argument('--headless')
    options.add_argument('--disable-dev-shm-usage')
    # options.add_argument('--proxy-server=93.157.248.108:88')
    # options.add_argument('--proxy-server=http://%s' % PROXY)
    try:
        driver = webdriver.Chrome(options)
    except (SessionNotCreatedException, WebDriverException):
        pass
        
    return driver


def get_used_categories():
    with open('used_categories.txt', 'r') as file:
        return [category.strip() for category in file.readlines()]


def get_list_of_category_ids():
    used_categories = get_used_categories()

    driver = get_web_driver()
    if not driver:
        return []

    driver.implicitly_wait(5)

    driver.get(url=f'{BASE_URL}')
    time.sleep(2)

    while True:
        try:

            menu_button = driver.find_element(By.CSS_SELECTOR, '[data-test="catalog-button"]').click()

            time.sleep(2)

            print('Получаем список категорий')
            list_of_categories = driver.find_element(By.CLASS_NAME, 'sections-menu-top')
            list_of_categories = list_of_categories.find_elements(By.TAG_NAME, 'ul')
            list_of_categories = list_of_categories[0].find_elements(By.TAG_NAME, 'li')
            break
        except (StaleElementReferenceException, NoSuchElementException):
            menu_button = driver.find_element(By.CSS_SELECTOR, '[data-test="catalog-button"]').click()

            time.sleep(1)

    action = ActionChains(driver)

    result = []
    count = 0

    for category in list_of_categories:
        action.move_to_element(category).perform()
        time.sleep(0.5)
        items = driver.find_elements(By.CLASS_NAME, 'masonry-brick')
        for item in items:
            if count >= 6:
                break

            item = item.find_element(By.CLASS_NAME, 'subsection').find_element(By.CLASS_NAME, 'subsection-childs')
            for child in item.find_elements(By.TAG_NAME, 'a'):
                category_href = child.get_attribute('href').split('/')[-2]
                if category_href not in used_categories:
                    result.append(category_href)
                    count += 1

                    if count >= 6:
                        break
        if count >= 6:
            break

    driver.close()

    return result


def get_page_info(category_id):
    driver = get_web_driver()
    if not driver:
        return {'category_title': '', 'items_count': 0, 'page_count': 0}
    driver.implicitly_wait(5)
    url = f'{BASE_URL}/catalog/{category_id}/'
    driver.get(url)

    tries_count = 0

    while True:
        if tries_count >= 5:
            driver.close()

            return {'category_title': '', 'items_count': 0, 'page_count': 0}
        try:
            items_count_elem = driver.find_element(By.CSS_SELECTOR, '[data-test="products-counter"]')
            items_count = int(items_count_elem.text.split(': ')[-1])

            break
        except (StaleElementReferenceException, NoSuchElementException):
            time.sleep(2)
            tries_count += 1


    # try:
    #     items_count_elem = driver.find_element(By.CSS_SELECTOR, '[data-test="products-counter"]')
    # except StaleElementReferenceException:
    #     time.sleep(1)
    #     items_count_elem = driver.find_element(By.CSS_SELECTOR, '[data-test="products-counter"]')

    # items_count_elem = driver.find_element(By.CSS_SELECTOR, '[data-test="products-counter"]')

    page_count = ceil(items_count / ITEMS_PER_PAGE) + 1

    while True:
        if tries_count >= 5:
            driver.close()

            return {'category_title': '', 'items_count': 0, 'page_count': 0}
        try:
            category_title = driver.find_element(By.CLASS_NAME, 'categories-title').text
            break
        except (StaleElementReferenceException, NoSuchElementException):
            time.sleep(2)
            tries_count += 1

    driver.close()

    return {'category_title': category_title, 'items_count': items_count, 'page_count': page_count}


def get_items_ids(category_id, max_pages):
    result = []

    for page in range(max_pages):
        driver = get_web_driver()
        if not driver:
            continue

        driver.get(f'{BASE_URL}/catalog/{category_id}/?p={page}')

        time.sleep(1.5)

        bs = BeautifulSoup(driver.page_source, 'html.parser')

        items = bs.find_all('div', class_='fade-in-list')

        for ind, item in enumerate(items, 1):
            while True:
                try:
                    item_info = item.find('p', class_='swiper-no-swiping')
                    result.append(item_info.text)
                    break
                except AttributeError:
                    print(f'{BASE_URL}/catalog/{category_id}/?p={page}')
                    time.sleep(1)
                    continue

        driver.close()

    result = list(set(result))

    return result


def get_category_info(category_id):
    page_info = get_page_info(category_id)
    if
    item_ids = get_items_ids(category_id, page_info['page_count'])

    return {'category_title': page_info['category_title'], 'category_id': category_id, 'item_ids': item_ids}


def get_item_data(item_id):
    driver = get_web_driver()
    if not driver:
        return {}


    # print(f'{BASE_URL}/product/{item_id}/')
    try:
        driver.get(f'{BASE_URL}/product/{item_id}/')
    except WebDriverException:
        driver.close()

        return {}

    time.sleep(0.5)

    tries_count = 0

    while True:
        if tries_count >= 10:
            driver.close()

            return {}
        try:
            driver.find_element(By.CSS_SELECTOR, '[data-test="product-characteristics-tab"]').click()
            break
        except (StaleElementReferenceException, NoSuchElementException):
            time.sleep(2)
            tries_count += 1
        except ElementClickInterceptedException:
            break
        except WebDriverException:
            driver.close()

            return {}

    bs = BeautifulSoup(driver.page_source, 'html.parser')

    driver.close()

    title = bs.find('h1', class_='title-lg')
    if title:
        title = title.text

    price = bs.find('p', attrs={'data-test': 'product-gold-price'})
    if price:
        price = price.text.split(r'\u')[0]

    item_data = {
        'ID': item_id,
        'Название': title,
        'Цена': price,
        'Ссылка': f'{BASE_URL}/product/{item_id}/',
    }

    product_tabs = bs.find_all('span', class_='product-title-text')

    properties_table = None

    for tab in product_tabs:
        if tab.text == 'Характеристики':
            properties_table = tab.parent.parent.find('ul', class_='product-properties-list listing-data')
            break

    if properties_table:
        for li in properties_table.find_all('li'):
            while True:
                try:
                    item_key = li.find('div', class_='title').text
                    item_value = li.find('div', class_='value').text
                    item_data[item_key] = item_value
                    break
                except AttributeError:
                    time.sleep(1)
                    continue

            # item_key = li.find('div', class_='title').text
            # item_value = li.find('div', class_='value').text

    return item_data


def upload_products_info(category_title, data):
    # data = [
    #     {'ID': '656216',
    #      'Название': 'Наличник 70х8х2150 мм финишпленка серый (1 шт.)',
    #      'Цена': '270 ₽',
    #      'Тип товара': 'Наличник',
    #      'Бренд': 'Verda ДПГ',
    #      'Страна-производитель': 'Россия',
    #      'Материал': 'МДФ',
    #      'Покрытие': 'Финишпленка',
    #      'Фактура': 'Гладкая',
    #      'Влагостойкость': 'Да',
    #      'Цвет производителя': 'Серый',
    #      'Длина, мм': '2150',
    #      'Форма': 'Плоская',
    #      'Конструкция наличника': 'Простая',
    #      'Коллекция': 'Verda',
    #      'Вес, кг': '0,8'},
    #     {'ID': '612180',
    #      'Название': 'Наличник гладкий 50х11х2200 мм сорт Экстра сращенный',
    #      'Цена': '198 ₽',
    #      'Артикул': '3802',
    #      'Тип товара': 'Наличник',
    #      'Сортность': 'Экстра',
    #      'Страна-производитель': 'Россия',
    #      'Толщина, мм': '11',
    #      'Материал': 'Массив хвойных пород',
    #      'Покрытие': 'Без покрытия',
    #      'Фактура': 'Гладкая',
    #      'Влагостойкость': 'Да',
    #      'Ширина, мм': '50',
    #      'Длина, мм': '2200',
    #      'Форма': 'Плоская',
    #      'Конструкция наличника': 'Простая',
    #      'Вес, кг': '0,61'},
    #     {'ID': '692289',
    #      'Название': 'Наличник Палитра/Сафари/Норд 70х8х2140 мм финишпленка белый '
    #                  'прямой (1 шт.)',
    #      'Цена': '260 ₽',
    #      'Тип товара': 'Наличник',
    #      'Бренд': 'СД',
    #      'Страна-производитель': 'Россия',
    #      'Материал': 'МДФ',
    #      'Покрытие': 'Финишпленка',
    #      'Фактура': 'Структурированная',
    #      'Влагостойкость': 'Да',
    #      'Цвет производителя': 'Белый',
    #      'Длина, мм': '2140',
    #      'Форма': 'Плоская',
    #      'Конструкция наличника': 'Простая',
    #      'Коллекция': 'Палитра',
    #      'Упаковка': 'Полиэтилен',
    #      'Вес, кг': '0,7'}
    # ]

    data = sorted(data, key=lambda x: x.get('Тип товара', ''))

    file = download_file_from_google_drive(settings.FILE_ID, settings.FILE_NAME)

    wb = load_workbook(settings.FILE_NAME)

    headers = []

    sheet_name = category_title
    if len(sheet_name) > 25:
        sheet_name = sheet_name[:25] + '...'

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    ws = wb[sheet_name]

    for product in data:
        for key in product.keys():
            if key not in headers:
                headers.append(key)

    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = value
        cell.alignment = align
        cell.border = bold_border
        cell.font = bold_font

    for ind, product in enumerate(data, 2):
        # product_type = product.get('Тип товара')
        if not product:
            continue

        max_row = ws.max_row + 1

        for row in ws.iter_rows(min_row=max_row, max_row=max_row, max_col=len(headers)):
            for col in row:
                col.border = thin_border

        for key, value in product.items():
            column = headers.index(key) + 1
            cell = ws.cell(row=max_row, column=column)
            cell.value = value

            if key == 'Ссылка':
                cell.hyperlink = value

    wb.save(settings.FILE_NAME)

    upload_file_to_google_drive(file, settings.FILE_NAME)
